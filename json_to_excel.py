#!/usr/bin/env python3
"""
BAC Checker v2.0 - JSON to Excel Converter
Converts test results JSON to color-coded Excel matrix

Color Coding:
- Green (FF00B050): 200 OK without redirect
- Orange (FFFFC000): Redirect (200→ or 301/302)
- Red (FFFF0000): Access Denied (403/404/401)
- Purple (FF7030A0): Server Error (500+)
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Color definitions (matching ACL_sample.xlsx)
COLORS = {
    'green': 'FF00B050',    # 200 OK (no redirect)
    'orange': 'FFFFC000',   # Redirect (200→ or 301/302)
    'red': 'FFFF0000',      # Access Denied (403/404/401)
    'purple': 'FF7030A0',   # Server Error (500+)
    'gray': 'FFD3D3D3',     # Error/Unknown
    'header': 'FF4472C4'    # Header background
}


def get_status_color(status_code):
    """
    Determine cell color based on status code.

    Args:
        status_code: HTTP status code string (e.g., "200", "200 →", "403")

    Returns:
        str: Hex color code
    """
    # Check if redirected (contains →)
    if '→' in status_code:
        return COLORS['orange']

    # Extract numeric status code
    code = status_code.strip().split()[0]

    try:
        code_int = int(code)
    except (ValueError, IndexError):
        return COLORS['gray']

    # Color mapping
    if code_int == 200:
        return COLORS['green']  # 200 without redirect
    elif 300 <= code_int < 400:
        return COLORS['orange']  # Redirects
    elif code_int in [401, 403, 404]:
        return COLORS['red']  # Access Denied
    elif 500 <= code_int < 600:
        return COLORS['purple']  # Server Error
    elif 200 <= code_int < 300:
        return COLORS['green']  # Other 2xx
    else:
        return COLORS['gray']  # Unknown


def json_to_excel(json_file, output_file=None):
    """
    Convert JSON results to color-coded Excel file.

    Args:
        json_file: Path to JSON results file
        output_file: Optional output Excel file path (auto-generated if None)

    Returns:
        str: Path to created Excel file
    """
    print("="*70)
    print("BAC Checker v2.0 - JSON to Excel Converter")
    print("="*70)

    # Load JSON data
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    urls = data.get('urls', [])
    roles = data.get('roles', [])
    results = data.get('results', {})

    print(f"Loaded: {len(urls)} URLs, {len(roles)} roles")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "BAC Test Results"

    # Define styles
    header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')

    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header row
    ws.cell(row=1, column=1, value='Paths')
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = header_alignment
    ws.cell(row=1, column=1).border = border_thin

    for col_idx, role in enumerate(roles, start=2):
        cell = ws.cell(row=1, column=col_idx, value=role)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_thin

    # Write data rows
    for row_idx, url in enumerate(urls, start=2):
        # Path column
        path_cell = ws.cell(row=row_idx, column=1, value=url)
        path_cell.border = border_thin
        path_cell.alignment = Alignment(horizontal='left', vertical='center')

        # Role columns
        for col_idx, role in enumerate(roles, start=2):
            status = results.get(url, {}).get(role, '000')

            cell = ws.cell(row=row_idx, column=col_idx, value=status)

            # Apply color based on status code
            color = get_status_color(status)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

            # Center alignment
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Border
            cell.border = border_thin

            # White text for better readability
            cell.font = Font(color='FFFFFFFF', bold=True, size=10)

    # Adjust column widths
    ws.column_dimensions['A'].width = 40  # Paths column
    for col_idx in range(2, len(roles) + 2):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 15

    # Add metadata sheet
    meta_ws = wb.create_sheet("Test Metadata")
    meta_ws.cell(row=1, column=1, value='Test Date')
    meta_ws.cell(row=1, column=2, value=data.get('test_date', 'Unknown'))

    meta_ws.cell(row=2, column=1, value='Total URLs')
    meta_ws.cell(row=2, column=2, value=len(urls))

    meta_ws.cell(row=3, column=1, value='Total Roles')
    meta_ws.cell(row=3, column=2, value=len(roles))

    meta_ws.cell(row=4, column=1, value='Total Tests')
    meta_ws.cell(row=4, column=2, value=len(urls) * len(roles))

    if data.get('stopped'):
        meta_ws.cell(row=5, column=1, value='Status')
        meta_ws.cell(row=5, column=2, value='STOPPED (Partial Results)')

    # Color legend
    meta_ws.cell(row=7, column=1, value='Color Legend')
    meta_ws.cell(row=7, column=1).font = Font(bold=True)

    legend_items = [
        ('Green', '200 OK (no redirect)', COLORS['green']),
        ('Orange', 'Redirect (200→ or 301/302)', COLORS['orange']),
        ('Red', 'Access Denied (403/404/401)', COLORS['red']),
        ('Purple', 'Server Error (500+)', COLORS['purple']),
        ('Gray', 'Error/Unknown', COLORS['gray'])
    ]

    for idx, (name, description, color) in enumerate(legend_items, start=8):
        meta_ws.cell(row=idx, column=1, value=name)
        meta_ws.cell(row=idx, column=2, value=description)
        meta_ws.cell(row=idx, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        meta_ws.cell(row=idx, column=1).font = Font(color='FFFFFFFF', bold=True)

    meta_ws.column_dimensions['A'].width = 15
    meta_ws.column_dimensions['B'].width = 40

    # Determine output file path
    if output_file is None:
        json_path = Path(json_file)
        output_file = json_path.parent / f"{json_path.stem}.xlsx"

    # Save workbook
    wb.save(output_file)

    print(f"✅ Excel file created: {output_file}")
    print("="*70)

    return str(output_file)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python json_to_excel.py <json_file> [output_file]")
        print("\nExample:")
        print("  python json_to_excel.py results/test_20251210_153045.json")
        print("  python json_to_excel.py results/test_20251210_153045.json custom_output.xlsx")
        sys.exit(1)

    json_file = sys.argv[1]

    if not Path(json_file).exists():
        print(f"ERROR: File not found: {json_file}")
        sys.exit(1)

    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        excel_file = json_to_excel(json_file, output_file)
        print(f"\n✅ Success! Open the file: {excel_file}")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
